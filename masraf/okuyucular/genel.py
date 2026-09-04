"""Genel amacli dosya okuma yardimcilari ve taninmayan dosyalar icin parser.

Bu modul okuyucular paketinin EN ALT katmanidir; paket icindeki hicbir
modulu import etmez (dairesel bagimlilik yok). antik.py, energo.py ve
kesif.py buradaki yardimcilari kullanir.

Iki is yapar:
    1. Excel (.xls / .xlsx / .xlsm) ve CSV dosyalarini tek bir sade
       "sayfa adi -> satir listesi" yapisina indirger (calisma_oku).
    2. Sablonunu tanimadigimiz dosyalar icin baslik satirini ve kolon
       anlamlarini anahtar kelimeyle tahmin eden bir yedek parser sunar
       (genel_oku).
"""

from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

from masraf.metin import ascii_katla
from masraf.modeller import GiderSatiri

__all__ = [
    "Calisma",
    "calisma_oku",
    "kolon_anahtari",
    "kolon_haritasi",
    "kolon_ara",
    "sayfa_sec",
    "baslik_satiri_bul",
    "hucre_metni",
    "hucre_sayisi",
    "hucre_tarihi",
    "tckn_normalize",
    "dolu_hucre_sayisi",
    "genel_oku",
]

# Bos sayilan hucre metinleri.
_BOS_METINLER = frozenset({"", "nan", "nat", "none", "null", "-", "#n/a", "na"})

# Excel seri tarih araligi (1970-01-01 ~ 25569, 2100-01-01 ~ 73051).
_SERI_ALT = 1.0
_SERI_UST = 80000.0

_TARIH_BICIMLERI: tuple[str, ...] = (
    "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y",
    "%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d",
    "%d.%m.%y", "%d/%m/%y",
    "%d.%m.%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S",
)

_SADECE_RAKAM = re.compile(r"\D+")
_COKLU_BOSLUK = re.compile(r"\s+")


# --------------------------------------------------------------------------
# Hucre degeri donusturucleri
# --------------------------------------------------------------------------

def hucre_metni(deger: Any) -> str | None:
    """Hucre degerini duz metne cevirir; anlamli icerik yoksa None doner.

    Bolunemez bosluk (\\xa0) ve coklu bosluklar tek boslugu indirgenir.
    """
    if deger is None:
        return None
    if isinstance(deger, str):
        metin = deger
    elif isinstance(deger, (datetime, date)):
        return deger.isoformat()
    elif isinstance(deger, float):
        # 4989.0 gibi degerleri '4989' olarak yaz, ondalikli ise oldugu gibi birak.
        if deger != deger:  # NaN
            return None
        metin = str(int(deger)) if deger.is_integer() else str(deger)
    else:
        metin = str(deger)
    metin = metin.replace("\xa0", " ").replace("​", "")
    metin = _COKLU_BOSLUK.sub(" ", metin).strip()
    if metin.lower() in _BOS_METINLER:
        return None
    return metin


def hucre_sayisi(deger: Any) -> float | None:
    """Hucre degerini ondalik sayiya cevirir; cevrilemezse None.

    Turkce ('1.234,56') ve Ingilizce ('1,234.56') bicimlerini ayirt eder,
    para birimi sembollerini ve bosluklari atar.
    """
    if deger is None or isinstance(deger, bool):
        return None
    if isinstance(deger, (int, float)):
        return None if deger != deger else float(deger)
    metin = hucre_metni(deger)
    if metin is None:
        return None
    metin = re.sub(r"[^\d,.\-]", "", metin)
    if not metin or metin in {"-", ".", ","}:
        return None
    son_nokta = metin.rfind(".")
    son_virgul = metin.rfind(",")
    if son_virgul > son_nokta:
        # Turkce bicim: nokta binlik ayiraci, virgul ondalik
        metin = metin.replace(".", "").replace(",", ".")
    else:
        # Ingilizce bicim: virgul binlik ayiraci
        metin = metin.replace(",", "")
    try:
        return float(metin)
    except ValueError:
        return None


def hucre_tarihi(deger: Any, datemode: int = 0) -> date | None:
    """Hucre degerini tarihe cevirir; cevrilemezse None.

    datetime/date nesnelerini, Excel seri numaralarini (xlrd datemode ile)
    ve yaygin metin bicimlerini ('22.05.2026', '2026-05-22') destekler.
    """
    if deger is None or isinstance(deger, bool):
        return None
    if isinstance(deger, datetime):
        return deger.date()
    if isinstance(deger, date):
        return deger
    if isinstance(deger, (int, float)):
        if deger != deger or not (_SERI_ALT <= float(deger) <= _SERI_UST):
            return None
        try:
            import xlrd

            return xlrd.xldate_as_datetime(float(deger), datemode).date()
        except Exception:
            return None
    metin = hucre_metni(deger)
    if metin is None:
        return None
    for bicim in _TARIH_BICIMLERI:
        try:
            return datetime.strptime(metin, bicim).date()
        except ValueError:
            continue
    # Salt sayi iceren metin Excel seri numarasi olabilir
    try:
        return hucre_tarihi(float(metin.replace(",", ".")), datemode)
    except ValueError:
        return None


def tckn_normalize(deger: Any) -> str | None:
    """TC Kimlik Numarasini dogrular ve 11 haneli metin olarak dondurur.

    Rakam disi karakterler atilir. Sonuc tam 11 hane degilse veya '0' ile
    basliyorsa None doner (gecersiz TCKN).
    """
    metin = hucre_metni(deger)
    if metin is None:
        return None
    if isinstance(deger, float) and deger.is_integer():
        metin = str(int(deger))
    rakamlar = _SADECE_RAKAM.sub("", metin)
    if len(rakamlar) != 11 or rakamlar[0] == "0":
        return None
    return rakamlar


def dolu_hucre_sayisi(satir: Sequence[Any]) -> int:
    """Satirdaki anlamli (bos olmayan) hucre sayisi."""
    return sum(1 for h in satir if hucre_metni(h) is not None)


def _metinsel_mi(deger: Any) -> bool:
    """Hucre saf metin mi (sayi veya tarih olarak yorumlanamiyor mu)?

    Baslik satirini veri satirlarindan ayirmak icin kullanilir: basliklar
    metin, veri satirlari genellikle sayi/tarih icerir. CSV'de her hucre
    metin olarak geldigi icin '15.07.2026' gibi degerler de veri sayilir.
    """
    if hucre_metni(deger) is None:
        return False
    return hucre_sayisi(deger) is None and hucre_tarihi(deger) is None


# --------------------------------------------------------------------------
# Kolon adi cozumleme
# --------------------------------------------------------------------------

def kolon_anahtari(ad: Any) -> str:
    """Kolon veya sayfa adini karsilastirilabilir ASCII/kucuk bicime cevirir.

    kayit._kolon_anahtari ile ayni davranisi gosterir; bolunemez bosluk,
    satir sonu, alt cizgi ve ayirici isaretler tek boslugu indirgenir.

    >>> kolon_anahtari("ADI SOYADI\\xa0\\xa0")
    'adi soyadi'
    """
    if ad is None:
        return ""
    metin = str(ad).replace("\xa0", " ")
    metin = unicodedata.normalize("NFKC", metin)
    metin = ascii_katla(metin)
    for isaret in ("\n", "\r", "\t", "_", "/", "-", ".", "(", ")", ",", ";", ":"):
        metin = metin.replace(isaret, " ")
    return " ".join(metin.lower().split())


def kolon_haritasi(baslik: Sequence[Any]) -> dict[str, int]:
    """Baslik satirindan 'normalize kolon adi -> indeks' haritasi uretir.

    Ayni ada sahip birden fazla kolon varsa ILK gecen kazanir.
    """
    harita: dict[str, int] = {}
    for i, hucre in enumerate(baslik):
        anahtar = kolon_anahtari(hucre)
        if anahtar and anahtar not in harita:
            harita[anahtar] = i
    return harita


def kolon_ara(
    harita: dict[str, int],
    *adaylar: str,
    icerir: bool = True,
    haric: Sequence[str] = (),
) -> int | None:
    """Aday adlardan biriyle eslesen kolonun indeksini dondurur.

    Once tam eslesme, sonra (icerir=True ise) 'aday, kolon adinin icinde
    geciyor mu' kontrolu yapilir. Adaylar oncelik sirasindadir.

    Args:
        haric: bu parcalari iceren kolon adlari hicbir zaman secilmez
            (orn. tarih ararken 'dogum tarihi' kolonunu elemek icin).
    """
    yasakli = [kolon_anahtari(h) for h in haric if kolon_anahtari(h)]

    def uygun(ad: str) -> bool:
        return not any(y in ad for y in yasakli)

    for aday in adaylar:
        anahtar = kolon_anahtari(aday)
        if anahtar in harita and uygun(anahtar):
            return harita[anahtar]
    if not icerir:
        return None
    for aday in adaylar:
        anahtar = kolon_anahtari(aday)
        if not anahtar:
            continue
        for ad, i in harita.items():
            if anahtar in ad and uygun(ad):
                return i
    return None


def sayfa_sec(sayfa_adlari: Iterable[str], *adaylar: str) -> str | None:
    """Sayfa adlarindan aday anahtarlarla eslesenin gercek adini dondurur.

    Turkce karakter farklarini yok saymak icin ascii katlanmis karsilastirma
    yapar ('Kisi Listesi' <-> 'Kişi Listesi').
    """
    adlar = list(sayfa_adlari)
    anahtarlar = {ad: kolon_anahtari(ad) for ad in adlar}
    for aday in adaylar:
        hedef = kolon_anahtari(aday)
        for ad in adlar:
            if anahtarlar[ad] == hedef:
                return ad
    for aday in adaylar:
        hedef = kolon_anahtari(aday)
        if not hedef:
            continue
        for ad in adlar:
            if hedef in anahtarlar[ad]:
                return ad
    return None


def baslik_satiri_bul(
    satirlar: Sequence[Sequence[Any]],
    aranan: Sequence[str] = (),
    sinir: int = 15,
) -> int:
    """Baslik satirinin 0 tabanli indeksini bulur.

    'aranan' verilmisse, ilk 'sinir' satir icinde bu anahtarlardan HERHANGI
    birini iceren ilk satir dondurulur (dinamik baslik tespiti). Bulunamazsa
    veya 'aranan' bos ise en cok dolu hucreye sahip ilk satir secilir.
    Hicbir dolu satir yoksa -1 doner.
    """
    ust = min(len(satirlar), sinir)
    if aranan:
        hedefler = {kolon_anahtari(a) for a in aranan if kolon_anahtari(a)}
        for i in range(ust):
            anahtarlar = {kolon_anahtari(h) for h in satirlar[i]}
            if hedefler & anahtarlar:
                return i
    # Puanlama: metin hucreleri baslik lehine, sayi/tarih hucreleri aleyhine.
    # Boylece 'veri satiri basliktan daha dolu' oldugu dosyalarda (orn. Yuzyil
    # dagitilmis) yanlislikla veri satiri baslik sanilmaz.
    en_iyi_i, en_iyi_puan = -1, 0.0
    for i in range(ust):
        satir = satirlar[i]
        metinsel = sum(1 for h in satir if _metinsel_mi(h))
        veri = dolu_hucre_sayisi(satir) - metinsel
        if metinsel < 2:
            continue
        puan = metinsel - 0.5 * veri
        if puan > en_iyi_puan:
            en_iyi_i, en_iyi_puan = i, puan
    return en_iyi_i


# --------------------------------------------------------------------------
# Dosya okuma
# --------------------------------------------------------------------------

@dataclass
class Calisma:
    """Bir Excel/CSV dosyasinin sade temsili.

    sayfalar: sayfa adi -> satir listesi (her satir hucre degerleri listesi).
    datemode: sadece .xls dosyalarinda anlamli (xlrd tarih taban modu).
    """

    yol: Path
    sayfalar: dict[str, list[list[Any]]] = field(default_factory=dict)
    datemode: int = 0

    @property
    def sayfa_adlari(self) -> list[str]:
        return list(self.sayfalar.keys())

    def satirlar(self, sayfa: str | None = None) -> list[list[Any]]:
        """Verilen sayfanin satirlarini dondurur; sayfa None ise ilk sayfa."""
        if not self.sayfalar:
            return []
        if sayfa is None:
            return next(iter(self.sayfalar.values()))
        return self.sayfalar.get(sayfa, [])


def _xls_oku(yol: Path, satir_siniri: int | None) -> Calisma:
    import xlrd

    kitap = xlrd.open_workbook(str(yol), formatting_info=False)
    try:
        sayfalar: dict[str, list[list[Any]]] = {}
        for sayfa in kitap.sheets():
            ust = sayfa.nrows if satir_siniri is None else min(sayfa.nrows, satir_siniri)
            sayfalar[sayfa.name] = [
                [sayfa.cell_value(r, c) for c in range(sayfa.ncols)] for r in range(ust)
            ]
        return Calisma(yol=yol, sayfalar=sayfalar, datemode=kitap.datemode)
    finally:
        kitap.release_resources()


def _xlsx_oku(yol: Path, satir_siniri: int | None) -> Calisma:
    import openpyxl

    kitap = openpyxl.load_workbook(str(yol), data_only=True, read_only=True)
    try:
        sayfalar: dict[str, list[list[Any]]] = {}
        for ad in kitap.sheetnames:
            sayfa = kitap[ad]
            satirlar: list[list[Any]] = []
            for i, satir in enumerate(sayfa.iter_rows(values_only=True)):
                if satir_siniri is not None and i >= satir_siniri:
                    break
                satirlar.append(list(satir))
            sayfalar[ad] = satirlar
        return Calisma(yol=yol, sayfalar=sayfalar, datemode=0)
    finally:
        kitap.close()


def _csv_oku(yol: Path, satir_siniri: int | None) -> Calisma:
    ham: str | None = None
    for kodlama in ("utf-8-sig", "cp1254", "latin-1"):
        try:
            ham = yol.read_text(encoding=kodlama)
            break
        except UnicodeDecodeError:
            continue
    if ham is None:
        ham = yol.read_text(encoding="utf-8", errors="replace")
    ornek = ham[:8192]
    try:
        lehce = csv.Sniffer().sniff(ornek, delimiters=";,\t|")
        ayirici = lehce.delimiter
    except csv.Error:
        ayirici = ";" if ornek.count(";") > ornek.count(",") else ","
    satirlar: list[list[Any]] = []
    for i, satir in enumerate(csv.reader(ham.splitlines(), delimiter=ayirici)):
        if satir_siniri is not None and i >= satir_siniri:
            break
        satirlar.append(list(satir))
    return Calisma(yol=yol, sayfalar={yol.stem: satirlar}, datemode=0)


def calisma_oku(yol: str | Path, satir_siniri: int | None = None) -> Calisma:
    """Excel veya CSV dosyasini sayfa adi -> satir listesi yapisina cevirir.

    satir_siniri verilirse her sayfadan sadece o kadar satir okunur; dosya
    tipi tespiti (kesif modulu) icin hizli on izleme saglar.

    Raises:
        FileNotFoundError: dosya yoksa.
        ValueError: uzanti desteklenmiyorsa.
    """
    p = Path(yol)
    if not p.exists():
        raise FileNotFoundError(f"Dosya bulunamadi: {p}")
    uzanti = p.suffix.lower()
    if uzanti == ".xls":
        return _xls_oku(p, satir_siniri)
    if uzanti in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _xlsx_oku(p, satir_siniri)
    if uzanti in {".csv", ".txt", ".tsv"}:
        return _csv_oku(p, satir_siniri)
    raise ValueError(f"Desteklenmeyen dosya uzantisi: {p.suffix} ({p.name})")


# --------------------------------------------------------------------------
# Taninmayan dosyalar icin yedek parser
# --------------------------------------------------------------------------

# Kolon anlami -> aday anahtar kelimeler (oncelik sirasinda).
_ISIM_ADAYLARI = (
    "ad soyad", "adi soyadi", "ad ve soyad", "isim soyisim", "isim",
    "personel adi", "personel", "katilimci", "calisan", "yolcu",
    "full name", "name", "aciklama",
)
_SICIL_ADAYLARI = (
    "sicil", "sicil no", "personel no", "personel numarasi", "id",
    "employee id", "tabel", "kod",
)
_TCKN_ADAYLARI = (
    "tckn", "tc kimlik no", "tc kimlik", "personel t c", "t c kimlik",
    "kimlik no", "tc no", "tc",
)
_TUTAR_ADAYLARI = (
    "tutar", "satis", "borc", "amount", "toplam", "bedel", "fiyat",
    "energo payi", "usd", "total",
)
_TARIH_ADAYLARI = (
    "belge tarihi", "fatura tarihi", "islem tarihi", "kayit tarihi",
    "tarih", "date",
)
_MERKEZ_ADAYLARI = (
    "santiye", "santiyesi", "proje", "masraf merkezi", "masraf yeri",
    "gorev yeri", "cost center", "yansitma", "ilgili sirket", "sirket",
)

# Tarih kolonu ararken asla secilmemesi gereken kolonlar.
_TARIH_HARIC = ("dogum tarihi", "dogum")

# Bu oneklerle baslayan 'isim' degerleri veri degil ozet satiridir.
_OZET_ONEKLERI: tuple[str, ...] = (
    "toplam", "genel toplam", "ara toplam", "odenecek", "iade", "total",
    "satir etiketleri", "genel", "ozet",
)

# Dosya adi / aciklama anahtar kelimesinden gider tipi tahmini.
_TIP_IPUCLARI: tuple[tuple[str, str], ...] = (
    ("konaklama", "Otel"),
    ("otel", "Otel"),
    ("hotel", "Otel"),
    ("bilet", "Bilet"),
    ("ucus", "Bilet"),
    ("vize", "Vize"),
    ("bagaj", "Bagaj"),
    ("arabulucu", "Arabuluculuk"),
    ("saglik", "Saglik"),
    ("egitim", "Egitim"),
    ("assessment", "Egitim"),
    ("katilimci", "Egitim"),
)


def _gider_tipi_tahmin(*metinler: str | None) -> str:
    """Dosya adi ve aciklamadan gider tipini tahmin eder; bulunamazsa 'Diger'."""
    birlesik = kolon_anahtari(" ".join(m for m in metinler if m))
    for anahtar, tip in _TIP_IPUCLARI:
        if anahtar in birlesik:
            return tip
    return "Diger"


def genel_oku(yol: str | Path) -> list[GiderSatiri]:
    """Sablonu taninmayan Excel/CSV dosyasini en iyi cabayla ayristirir.

    Baslik satirini ilk 10 satir icinde en cok dolu hucreye sahip satir
    olarak belirler, ardindan kolon adlarindan isim / sicil / tckn / tutar /
    tarih / masraf merkezi kolonlarini anahtar kelimeyle tahmin eder.
    Bulunamayan alanlar None birakilir.

    Tum sayfalar taranir; her sayfa icin ayri baslik/kolon cozumu yapilir.
    """
    p = Path(yol)
    calisma = calisma_oku(p)
    sonuclar: list[GiderSatiri] = []

    for sayfa_adi, satirlar in calisma.sayfalar.items():
        baslik_i = baslik_satiri_bul(satirlar, sinir=10)
        if baslik_i < 0:
            continue
        harita = kolon_haritasi(satirlar[baslik_i])
        if not harita:
            continue

        # Kullanici sozlugu (veri/kolon_esanlamlilari.csv) yerlesik adaylarin
        # ONUNE eklenir. Boylece yeni bir tedarikci sablonu geldiginde kod
        # degistirmeden, tek satir CSV ile taninabilir hale gelir.
        try:
            from masraf.kolon_sozlugu import genislet as _genislet
        except Exception:  # noqa: BLE001
            def _genislet(alan, varsayilanlar, veri_dizini="veri"):
                return tuple(varsayilanlar)

        i_isim = kolon_ara(harita, *_genislet("kisi", _ISIM_ADAYLARI))
        i_sicil = kolon_ara(harita, *_genislet("sicil", _SICIL_ADAYLARI))
        i_tckn = kolon_ara(harita, *_genislet("tckn", _TCKN_ADAYLARI))
        i_tutar = kolon_ara(harita, *_genislet("tutar", _TUTAR_ADAYLARI))
        i_tarih = kolon_ara(harita, *_genislet("tarih", _TARIH_ADAYLARI), haric=_TARIH_HARIC)
        i_merkez = kolon_ara(harita, *_genislet("santiye", _MERKEZ_ADAYLARI))

        # Ayni kolon hem isim hem masraf merkezi olarak secilmesin.
        if i_merkez is not None and i_merkez == i_isim:
            i_merkez = None
        if i_sicil is not None and i_sicil == i_tckn:
            i_sicil = None

        if i_isim is None and i_sicil is None and i_tckn is None:
            continue  # kisi bilgisi yok, bu sayfa gider satiri uretmez

        from masraf.kayit import sicil_normalize

        for r in range(baslik_i + 1, len(satirlar)):
            satir = satirlar[r]
            if dolu_hucre_sayisi(satir) == 0:
                continue

            def al(i: int | None) -> Any:
                if i is None or i >= len(satir):
                    return None
                return satir[i]

            isim = hucre_metni(al(i_isim))
            sicil = sicil_normalize(al(i_sicil)) or None
            tckn = tckn_normalize(al(i_tckn))
            if isim is None and sicil is None and tckn is None:
                continue  # bos satir
            if tckn is None:
                # TOPLAM / IADE / ODENECEK gibi ozet satirlarini ele.
                # TCKN varsa satir kesin bir kisiye aittir, elenmez.
                kimlikler = [kolon_anahtari(k) for k in (isim, sicil) if k]
                if any(
                    k.startswith(o) for k in kimlikler for o in _OZET_ONEKLERI
                ):
                    continue

            aciklama = isim or " | ".join(
                m for m in (hucre_metni(h) for h in satir) if m
            )
            sonuclar.append(
                GiderSatiri(
                    kaynak_dosya=p.name,
                    kaynak_tip="genel",
                    satir_no=r + 1,
                    belge_tarihi=hucre_tarihi(al(i_tarih), calisma.datemode),
                    aciklama=aciklama,
                    kisi_ham=isim,
                    sicil_ham=sicil,
                    tckn_ham=tckn,
                    tutar=hucre_sayisi(al(i_tutar)),
                    para_birimi=None,
                    masraf_merkezi_kaynak=hucre_metni(al(i_merkez)),
                    gider_tipi=_gider_tipi_tahmin(p.name, sayfa_adi),
                    ek={
                        "sayfa": sayfa_adi,
                        "baslik_satiri": baslik_i + 1,
                        "cozulen_kolonlar": {
                            "isim": i_isim, "sicil": i_sicil, "tckn": i_tckn,
                            "tutar": i_tutar, "tarih": i_tarih, "merkez": i_merkez,
                        },
                    },
                )
            )
    return sonuclar
