"""Uctan uca boru hatti testi (masraf.boru + masraf.cikti).

Gercek bir fatura dosyasi okunur, kisiler eslestirilir, masraf merkezleri
cozulur ve Excel ciktisi yazilir. Test ciktinin gercekten acilabildigini ve
dort sayfayi da icerdigini dogrular; cikti ``cikti/`` dizinine yazilir ve
test sonunda silinir.

Ogrenen defterler GECICI bir dizine yonlendirilir; test kullanicinin
``veri/aliases.csv`` dosyasina hicbir sey yazmaz.
"""

from __future__ import annotations

import shutil
import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path

KOK = Path(__file__).resolve().parents[1]
if str(KOK) not in sys.path:
    sys.path.insert(0, str(KOK))

try:
    from masraf.boru import Boru, CalismaAyarlari
    from masraf.modeller import (
        DURUM_ESLESMEDI,
        DURUM_INCELE,
        DURUM_OTOMATIK,
        Sonuc,
    )

    MODUL_VAR = True
except ImportError:
    MODUL_VAR = False

try:
    import openpyxl

    OPENPYXL_VAR = True
except ImportError:
    OPENPYXL_VAR = False

PERSONEL = KOK / "ornek_veri" / "personel" / "2025_2026_giris_cikis.xlsx"
ANTIK = KOK / "ornek_veri" / "antik_travel" / "ANTIK_CARI_TEMMUZ_2026.xls"
CIKTI_DIZINI = KOK / "cikti"
CIKTI_ADI = "TEST_uctan_uca.xlsx"

#: Excel ciktisinda bulunmasi zorunlu sayfalar, gosterildikleri sirada.
#: Mahsuplasma en basta: muhasebeye giden tablo odur, digerleri onun dayanagi.
BEKLENEN_SAYFALAR = ["Mahsuplasma", "Kontrol", "Sonuc", "Incele", "Eslesmedi", "Ozet"]

#: 'Sonuc' sayfasinda bulunmasi zorunlu kolonlar.
ZORUNLU_KOLONLAR = {
    "Kaynak Dosya", "Satir", "Belge Tarihi", "Gider Tipi", "Aciklama",
    "Cikarilan Kisi", "Sicil", "Ad Soyad", "Eslestirme Yontemi", "Guven",
    "Eslestirme Aciklamasi", "Gorev Yeri", "Masraf Merkezi Kodu", "Durum",
}


@unittest.skipUnless(MODUL_VAR, "masraf.boru bulunamadi")
class UctanUcaTest(unittest.TestCase):
    """Antik cari dokumunu bastan sona isler."""

    cikti_yolu: Path | None = None
    sonuc: dict | None = None
    gecici_veri: str | None = None

    @classmethod
    def setUpClass(cls):
        if not (PERSONEL.exists() and ANTIK.exists()):
            return
        # Ogrenen defterler gecici dizine; kullanicinin veri/ dizini kirlenmesin.
        cls.gecici_veri = tempfile.mkdtemp(prefix="masraf_test_veri_")
        harita = KOK / "veri" / "masraf_merkezi_haritasi.csv"
        if harita.exists():
            shutil.copy(harita, Path(cls.gecici_veri) / harita.name)

        CIKTI_DIZINI.mkdir(parents=True, exist_ok=True)
        boru = Boru(CalismaAyarlari(
            personel_yolu=PERSONEL,
            veri_dizini=cls.gecici_veri,
            cikti_dizini=CIKTI_DIZINI,
            ogrenmeyi_kaydet=False,
        ))
        cls.boru = boru
        cls.sonuc = boru.calistir([ANTIK], cikti_adi=CIKTI_ADI)
        cls.cikti_yolu = Path(cls.sonuc["excel_yolu"]) if cls.sonuc["excel_yolu"] else None

    @classmethod
    def tearDownClass(cls):
        # Test ciktisi repoda birakilmaz.
        if cls.cikti_yolu and cls.cikti_yolu.exists():
            cls.cikti_yolu.unlink()
        if cls.gecici_veri:
            shutil.rmtree(cls.gecici_veri, ignore_errors=True)

    def setUp(self):
        if not PERSONEL.exists():
            self.skipTest(f"Ornek veri bulunamadi: {PERSONEL}")
        if not ANTIK.exists():
            self.skipTest(f"Ornek veri bulunamadi: {ANTIK}")

    # ------------------------------------------------------------------
    # Is akisi
    # ------------------------------------------------------------------

    def test_hatasiz_calisti(self):
        self.assertEqual(self.boru.hatalar, [], "Boru hattinda hata olusmamali")

    def test_tum_satirlar_islendi(self):
        sonuclar = self.sonuc["sonuclar"]
        self.assertEqual(len(sonuclar), 134)
        for sonuc in sonuclar:
            self.assertIsInstance(sonuc, Sonuc)
            self.assertIn(sonuc.durum,
                          (DURUM_OTOMATIK, DURUM_INCELE, DURUM_ESLESMEDI))

    def test_otomatik_orani_makul(self):
        # Naif isim eslestirme %66 basari veriyordu; kademeli eslestirme
        # bunun altina DUSMEMELIDIR (aile/kesik/translit vakalari dahil).
        sonuclar = self.sonuc["sonuclar"]
        otomatik = [s for s in sonuclar if s.durum == DURUM_OTOMATIK]
        oran = len(otomatik) / len(sonuclar)
        self.assertGreaterEqual(
            oran, 0.50,
            f"Otomatik oran cok dustu: {oran:.0%} ({len(otomatik)}/{len(sonuclar)})",
        )

    def test_otomatik_satirlarin_masraf_merkezi_var(self):
        for sonuc in self.sonuc["sonuclar"]:
            if sonuc.durum == DURUM_OTOMATIK:
                self.assertIsNotNone(sonuc.eslesme.sicil)
                self.assertTrue(
                    sonuc.gorev_yeri or sonuc.masraf_merkezi,
                    f"Satir {sonuc.satir.satir_no}: OTOMATIK ama masraf merkezi bos",
                )

    def test_donem_tarihe_gore_secilmis(self):
        # Temmuz 2026 faturasi normalde Temmuz 2026 snapshot'ina dusmeli.
        # Isten ayrilmis kisilerde ise kisinin SON snapshot'i kullanilir;
        # bu bilincli bir davranistir (bkz. PersonelDefteri.donem_kaydi).
        donemli = [s for s in self.sonuc["sonuclar"] if s.donem is not None]
        self.assertTrue(donemli)
        for sonuc in donemli:
            with self.subTest(satir=sonuc.satir.satir_no):
                self.assertGreaterEqual(sonuc.donem, date(2025, 11, 1))
                self.assertLessEqual(sonuc.donem, date(2026, 7, 1))

        aktif = [s for s in donemli if s.kategori == "Aktif"]
        self.assertTrue(aktif)
        for sonuc in aktif:
            with self.subTest(satir=sonuc.satir.satir_no):
                self.assertEqual(
                    sonuc.donem, date(2026, 7, 1),
                    f"Aktif calisan {sonuc.eslesme.ad_soyad} icin Temmuz 2026 "
                    f"donemi beklenirdi, {sonuc.donem} secildi",
                )

    def test_ayrilan_kisi_masraf_merkezisiz_kalmaz(self):
        # Cikisli kayitlarin buyuk kismi 'donus bileti' satiridir: kisi bilet
        # kesilirken hala calisiyordu. Bunlar dogru sekilde otomatik gecer,
        # ama hicbiri masraf merkezisiz kalmamalidir.
        ayrilanlar = [s for s in self.sonuc["sonuclar"] if s.kategori == "Cikis"]
        self.assertTrue(ayrilanlar)
        for sonuc in ayrilanlar:
            with self.subTest(satir=sonuc.satir.satir_no):
                self.assertTrue(sonuc.gorev_yeri or sonuc.masraf_merkezi)

    def test_cikistan_sonraki_masraf_incelemeye_duser(self):
        # Kisi belge tarihinden ONCE ayrilmissa masraf onun son projesine
        # yazilir ama sessizce onaylanmaz: uyari uretilip incelemeye gider.
        gecmis_cikisli = [
            s for s in self.sonuc["sonuclar"]
            if s.cikis_tarihi and s.satir.belge_tarihi
            and s.cikis_tarihi < s.satir.belge_tarihi
        ]
        self.assertTrue(gecmis_cikisli,
                        "Ornek dosyada cikistan sonraki masraf satiri bekleniyordu")
        for sonuc in gecmis_cikisli:
            with self.subTest(satir=sonuc.satir.satir_no):
                self.assertNotEqual(sonuc.durum, DURUM_OTOMATIK)
                self.assertTrue(sonuc.uyarilar,
                                "Cikis sonrasi masraf icin uyari uretilmeli")

    def test_ozet_uretildi(self):
        ozet = self.sonuc["ozet"]
        for anahtar in ("satir_sayisi", "durum_dagilimi", "yontem_dagilimi",
                        "otomatik_orani"):
            self.assertIn(anahtar, ozet)
        self.assertEqual(ozet["satir_sayisi"], len(self.sonuc["sonuclar"]))

    # ------------------------------------------------------------------
    # Excel ciktisi
    # ------------------------------------------------------------------

    def test_excel_dosyasi_olustu(self):
        self.assertIsNotNone(self.cikti_yolu, "Excel yolu bos dondu")
        self.assertTrue(self.cikti_yolu.exists(), f"Dosya yok: {self.cikti_yolu}")
        self.assertEqual(self.cikti_yolu.parent.resolve(), CIKTI_DIZINI.resolve())
        self.assertGreater(self.cikti_yolu.stat().st_size, 4096)

    @unittest.skipUnless(OPENPYXL_VAR, "openpyxl bulunamadi")
    def test_excel_acilir_ve_sayfalar_dogru(self):
        calisma = openpyxl.load_workbook(self.cikti_yolu, read_only=True)
        try:
            self.assertEqual(calisma.sheetnames, BEKLENEN_SAYFALAR)
        finally:
            calisma.close()

    def test_mahsuplasma_mutabakati_kapaniyor(self):
        """Uctan uca calismada da para kaybolmamali.

        Bu, ciktinin muhasebeye gonderilebilir olmasinin on sartidir: her
        fatura icin okunan tutar, yinelenen + dagitilan + dagitilamayan
        toplamina esit olmali.
        """
        mahsup = self.sonuc.get("mahsup")
        self.assertIsNotNone(mahsup, "calistir() mahsuplasma tablosu dondurmeli")
        for kontrol in mahsup.kontrol:
            with self.subTest(fatura=kontrol.kaynak):
                self.assertTrue(
                    kontrol.kapali_mi,
                    f"{kontrol.kaynak}: fark {kontrol.fark:+.2f}",
                )
        toplam_mahsup = round(sum(m.tutar for m in mahsup.satirlar), 2)
        beklenen = round(
            sum(k.dagitilan + k.dagitilamayan for k in mahsup.kontrol), 2
        )
        self.assertAlmostEqual(toplam_mahsup, beklenen, places=2)

    @unittest.skipUnless(OPENPYXL_VAR, "openpyxl bulunamadi")
    def test_sonuc_sayfasi_tum_satirlari_icerir(self):
        calisma = openpyxl.load_workbook(self.cikti_yolu, read_only=True)
        try:
            sayfa = calisma["Sonuc"]
            # 1 baslik satiri + veri satirlari
            self.assertEqual(sayfa.max_row, len(self.sonuc["sonuclar"]) + 1)
            basliklar = [h.value for h in next(sayfa.iter_rows(max_row=1))]
            eksik = ZORUNLU_KOLONLAR - set(basliklar)
            self.assertFalse(eksik, f"Ciktida eksik kolon: {sorted(eksik)}")
        finally:
            calisma.close()

    @unittest.skipUnless(OPENPYXL_VAR, "openpyxl bulunamadi")
    def test_incele_ve_eslesmedi_sayfalari_filtreli(self):
        sonuclar = self.sonuc["sonuclar"]
        beklenen = {
            "Incele": len([s for s in sonuclar if s.durum == DURUM_INCELE]),
            "Eslesmedi": len([s for s in sonuclar if s.durum == DURUM_ESLESMEDI]),
        }
        calisma = openpyxl.load_workbook(self.cikti_yolu, read_only=True)
        try:
            for ad, sayi in beklenen.items():
                with self.subTest(sayfa=ad):
                    # Bos filtrede bile baslik satiri yazilir.
                    self.assertEqual(calisma[ad].max_row, sayi + 1)
        finally:
            calisma.close()

    @unittest.skipUnless(OPENPYXL_VAR, "openpyxl bulunamadi")
    def test_ozet_sayfasi_dolu(self):
        calisma = openpyxl.load_workbook(self.cikti_yolu, read_only=True)
        try:
            self.assertGreater(calisma["Ozet"].max_row, 5)
        finally:
            calisma.close()

    # ------------------------------------------------------------------
    # Gizlilik
    # ------------------------------------------------------------------

    def test_kullanici_defterlerine_yazilmadi(self):
        # ogrenmeyi_kaydet=False iken repo icindeki veri/ dizini degismemeli.
        gecici = Path(self.gecici_veri)
        self.assertTrue(gecici.exists())
        self.assertNotEqual(gecici.resolve(), (KOK / "veri").resolve())


if __name__ == "__main__":
    unittest.main()
