"""Mahsuplasma (dagitim) tablosu testleri.

Bu modulun tek bir sozu vardir: PARA KAYBOLMAZ. Her fatura icin

    okunan = yinelenen + dagitilan + dagitilamayan

kurusuna kadar tutmalidir. Testlerin cogu bu esitligi farkli acilardan zorlar.

Ikinci sozu: ayni islem iki dosyada gelirse BIR KEZ sayilir. Gercek Temmuz
2026 mailinde acentenin ham dokumu (48.946,59 USD) ile ayni islemlerin elle
dagitilmis hali (48.978,59 USD) birlikte geliyor. Naif toplama parayi ikiye
katlar.
"""

from __future__ import annotations

import unittest
from datetime import date
from pathlib import Path

from masraf.mahsuplasma import (
    DAGITILAMAYAN,
    mahsuplasma_uret,
)
from masraf.modeller import (
    DURUM_ESLESMEDI,
    DURUM_INCELE,
    DURUM_OTOMATIK,
    Eslesme,
    GiderSatiri,
    Sonuc,
)


def gider(
    kaynak: str = "fatura.xlsx",
    tip: str = "genel",
    satir_no: int = 1,
    tarih: date | None = date(2026, 7, 15),
    kisi: str | None = "AHMET YILMAZ",
    tutar: float | None = 100.0,
    para: str = "USD",
    gider_tipi: str = "Bilet",
    santiye: str | None = None,
    ek: dict | None = None,
) -> GiderSatiri:
    return GiderSatiri(
        kaynak_dosya=kaynak, kaynak_tip=tip, satir_no=satir_no,
        belge_tarihi=tarih, aciklama=kisi or "", kisi_ham=kisi,
        sicil_ham=None, tckn_ham=None, tutar=tutar, para_birimi=para,
        masraf_merkezi_kaynak=santiye, gider_tipi=gider_tipi, ek=dict(ek or {}),
    )


def sonuc(
    satir: GiderSatiri,
    merkez: str | None = "GPP",
    durum: str = DURUM_OTOMATIK,
    sicil: str | None = "100001",
    sirket: str | None = "RHI",
    haritada: bool = True,
) -> Sonuc:
    satir.ek.setdefault("masraf_merkezi_adi", "GPP Project" if merkez else "")
    satir.ek.setdefault("masraf_merkezi_haritada", haritada)
    return Sonuc(
        satir=satir,
        eslesme=Eslesme(sicil=sicil, ad_soyad=satir.kisi_ham, yontem="tam_isim",
                        guven=1.0, aday_sayisi=1, aciklama=""),
        donem=date(2026, 7, 1), gorev_yeri="Ust-Luga GPC",
        masraf_merkezi=merkez, sirket=sirket, sirket2=sirket,
        statu="Aktif", kategori="Aktif", cikis_tarihi=None, durum=durum,
    )


class ParaKaybolmazTest(unittest.TestCase):
    def test_tek_fatura_kapaniyor(self):
        s = [sonuc(gider(satir_no=i, tutar=t)) for i, t in enumerate([100.0, 250.5, 33.33], 1)]
        t = mahsuplasma_uret(s)
        self.assertTrue(t.kapali_mi, t.acik_kontroller)
        self.assertAlmostEqual(t.kontrol[0].gelen, 383.83, places=2)
        self.assertAlmostEqual(t.kontrol[0].dagitilan, 383.83, places=2)

    def test_ucte_bir_bolme_kurus_birakmaz(self):
        """100/3 oransal bolununce 33.33 x 3 = 99.99 eder. Artik dagitilmali."""
        paylasim = [
            {"masraf_merkezi": "A", "pay": 1, "bolen": 3, "oran": 1 / 3},
            {"masraf_merkezi": "B", "pay": 1, "bolen": 3, "oran": 1 / 3},
            {"masraf_merkezi": "C", "pay": 1, "bolen": 3, "oran": 1 / 3},
        ]
        t = mahsuplasma_uret([sonuc(gider(tutar=100.0, ek={"paylasim": paylasim}))])
        self.assertEqual(len(t.satirlar), 3)
        self.assertAlmostEqual(sum(m.tutar for m in t.satirlar), 100.0, places=2)
        self.assertTrue(t.kapali_mi)

    def test_dagitilamayan_gorunur_kalir(self):
        s = [sonuc(gider(satir_no=1, tutar=100.0)),
             sonuc(gider(satir_no=2, tutar=60.0), merkez=None,
                   durum=DURUM_ESLESMEDI, sicil=None)]
        t = mahsuplasma_uret(s)
        dagitilamayan = [m for m in t.satirlar if m.masraf_merkezi == DAGITILAMAYAN]
        self.assertEqual(len(dagitilamayan), 1)
        self.assertAlmostEqual(dagitilamayan[0].tutar, 60.0, places=2)
        self.assertAlmostEqual(t.kontrol[0].dagitilamayan, 60.0, places=2)
        self.assertTrue(t.kapali_mi)

    def test_negatif_tutar_dusulur(self):
        """Iade satiri toplami azaltmali, mutlak deger olarak eklenmemeli."""
        s = [sonuc(gider(satir_no=1, tutar=500.0)),
             sonuc(gider(satir_no=2, tutar=-120.0, kisi="MEHMET KAYA", tarih=date(2026, 7, 20)))]
        t = mahsuplasma_uret(s)
        self.assertAlmostEqual(t.kontrol[0].gelen, 380.0, places=2)
        self.assertAlmostEqual(t.kontrol[0].dagitilan, 380.0, places=2)
        self.assertTrue(t.kapali_mi)

    def test_para_birimleri_karismaz(self):
        s = [sonuc(gider(satir_no=1, tutar=100.0, para="USD")),
             sonuc(gider(satir_no=2, tutar=100.0, para="RUB", kisi="MEHMET KAYA"))]
        t = mahsuplasma_uret(s)
        self.assertEqual({k.para_birimi for k in t.kontrol}, {"USD", "RUB"})
        self.assertEqual(set(t.toplamlar()), {"USD", "RUB"})
        self.assertTrue(t.kapali_mi)

    def test_tutarsiz_ve_kutuk_satirlari_dagilima_girmez(self):
        s = [
            sonuc(gider(satir_no=1, tutar=100.0)),
            sonuc(gider(satir_no=2, tutar=None, kisi="TUTARSIZ KISI")),
            sonuc(gider(satir_no=3, tutar=None, tip="koc_katilimci", kisi="KATILIMCI")),
            sonuc(gider(satir_no=4, tutar=None, tip="energo_saglik", kisi="SAGLIK")),
        ]
        t = mahsuplasma_uret(s)
        self.assertEqual(t.kutuk_satir_sayisi, 2)
        self.assertEqual(t.tutarsiz_satir_sayisi, 1)
        self.assertAlmostEqual(t.kontrol[0].gelen, 100.0, places=2)


class YinelemeTest(unittest.TestCase):
    """Ayni islem iki dosyada: bir kez sayilmali."""

    def _cift(self, isim_a: str, isim_b: str, tutar_a=250.0, tutar_b=250.0):
        ham = sonuc(gider(kaynak="ENERGO.xls", tip="antik_cari",
                          kisi=isim_a, tutar=tutar_a))
        elle = sonuc(gider(kaynak="YUZYIL.xlsx", tip="yuzyil_dagitilmis",
                           kisi=isim_b, tutar=tutar_b))
        return mahsuplasma_uret([ham, elle])

    def test_ayni_isim_yinelenme_sayilir(self):
        t = self._cift("AHMET YILMAZ", "AHMET YILMAZ")
        self.assertEqual(t.yinelenen_sayisi, 1)
        self.assertAlmostEqual(t.toplamlar()["USD"]["net"], 250.0, places=2)

    def test_isim_sirasi_farkli_olsa_da_yakalanir(self):
        """Ham dokum 'OZAKAY MUSTAFAKEMAL', elle dagitilmis 'MUSTAFA KEMAL OZAKAY'."""
        t = self._cift("OZAKAY MUSTAFAKEMAL", "MUSTAFA KEMAL OZAKAY")
        self.assertEqual(t.yinelenen_sayisi, 1)

    def test_kirpilmis_isim_de_yakalanir(self):
        t = self._cift("OZAKAY MUSTAFAKEMAL", "OZAKAY MUSTAFAKEMA")
        self.assertEqual(t.yinelenen_sayisi, 1)

    def test_isimsiz_satirlar_da_yakalanir(self):
        """Kurumsal kalemlerde (celenk, toplanti) kisi yoktur; tarih+tutar yeter."""
        t = self._cift(None, None)
        self.assertEqual(t.yinelenen_sayisi, 1)

    def test_ham_dokum_tercih_edilir(self):
        t = self._cift("AHMET YILMAZ", "AHMET YILMAZ")
        kalan = [k for k in t.kontrol if k.yinelenen_satir == 0]
        self.assertEqual(kalan[0].kaynak, "ENERGO.xls")

    def test_isaret_celiskisi_raporlanir(self):
        """Ham dokumde -16, elle dagitilmis halde +16: gercek veride var."""
        t = self._cift("EROL GUNES", "EROL GUNES", tutar_a=-16.0, tutar_b=16.0)
        self.assertEqual(t.yinelenen_sayisi, 1)
        self.assertEqual(len(t.isaret_celiskileri), 1)
        self.assertAlmostEqual(t.isaret_celiskileri[0].kullanilan, -16.0, places=2)
        self.assertAlmostEqual(t.toplamlar()["USD"]["net"], -16.0, places=2)

    def test_ayni_dosyadaki_tekrarlar_korunur(self):
        """Ayni ucus, ayni ucret, FARKLI kisiler. Bunlar yineleme degildir."""
        s = [sonuc(gider(kaynak="ENERGO.xls", tip="antik_cari", satir_no=1,
                         kisi="AHMET YILMAZ", tutar=227.46)),
             sonuc(gider(kaynak="ENERGO.xls", tip="antik_cari", satir_no=2,
                         kisi="MEHMET KAYA", tutar=227.46))]
        t = mahsuplasma_uret(s)
        self.assertEqual(t.yinelenen_sayisi, 0)
        self.assertAlmostEqual(t.kontrol[0].gelen, 454.92, places=2)

    def test_fazlalik_satir_korunur(self):
        """Elle dagitilmis dosyada ham dokumde olmayan bir kalem varsa kaybolmaz."""
        ham = sonuc(gider(kaynak="ENERGO.xls", tip="antik_cari",
                          kisi="AHMET YILMAZ", tutar=250.0))
        elle1 = sonuc(gider(kaynak="YUZYIL.xlsx", tip="yuzyil_dagitilmis",
                            satir_no=1, kisi="AHMET YILMAZ", tutar=250.0))
        elle2 = sonuc(gider(kaynak="YUZYIL.xlsx", tip="yuzyil_dagitilmis",
                            satir_no=2, kisi="VELI DEMIR", tutar=250.0))
        t = mahsuplasma_uret([ham, elle1, elle2])
        self.assertEqual(t.yinelenen_sayisi, 1)
        self.assertAlmostEqual(t.toplamlar()["USD"]["net"], 500.0, places=2)
        self.assertTrue(t.kapali_mi)

    def test_yineleme_kapatilabilir(self):
        ham = sonuc(gider(kaynak="ENERGO.xls", tip="antik_cari", tutar=250.0))
        elle = sonuc(gider(kaynak="YUZYIL.xlsx", tip="yuzyil_dagitilmis", tutar=250.0))
        t = mahsuplasma_uret([ham, elle], yinelenenleri_ele=False)
        self.assertEqual(t.yinelenen_sayisi, 0)
        self.assertAlmostEqual(t.toplamlar()["USD"]["net"], 500.0, places=2)

    def test_dagitim_talimati_devralinir(self):
        """Paylasim karari sadece elle dagitilmis dosyada olabilir; kaybolmamali."""
        paylasim = [{"masraf_merkezi": "RHI", "pay": 1, "bolen": 3, "oran": 1 / 3},
                    {"masraf_merkezi": "RENSTROYDETAL", "pay": 2, "bolen": 3, "oran": 2 / 3}]
        ham = sonuc(gider(kaynak="ENERGO.xls", tip="antik_cari", tutar=201.42))
        elle = sonuc(gider(kaynak="YUZYIL.xlsx", tip="yuzyil_dagitilmis",
                           tutar=201.42, ek={"paylasim": paylasim}))
        t = mahsuplasma_uret([ham, elle])
        notlar = sorted(m.pay_notu for m in t.satirlar if m.pay_notu)
        self.assertEqual(notlar, ["RENSTROYDETAL 2/3", "RHI 1/3"])
        self.assertAlmostEqual(sum(m.tutar for m in t.satirlar), 201.42, places=2)


class PaylasimTest(unittest.TestCase):
    def test_tuzel_kisi_paylasimi_projeyi_bolmez(self):
        """'RHI 1/3 - RENSTROYDETAL 2/3' sirket paylasimidir, proje ayni kalir."""
        paylasim = [{"masraf_merkezi": "RHI", "pay": 1, "bolen": 3, "oran": 1 / 3},
                    {"masraf_merkezi": "RENSTROYDETAL", "pay": 2, "bolen": 3, "oran": 2 / 3}]
        t = mahsuplasma_uret([sonuc(gider(tutar=201.42, ek={"paylasim": paylasim}))])
        self.assertEqual(len(t.satirlar), 2)
        self.assertEqual({m.masraf_merkezi for m in t.satirlar}, {"GPP"})
        self.assertEqual({m.sirket for m in t.satirlar}, {"RHI", "RENSTROYDETAL"})
        buyuk = max(t.satirlar, key=lambda m: m.tutar)
        self.assertEqual(buyuk.sirket, "RENSTROYDETAL")
        self.assertAlmostEqual(buyuk.tutar, 134.28, places=2)

    def test_proje_paylasimi_haritayla_bolunur(self):
        """Etiket gercekten bir projeyse masraf merkezi de bolunur."""

        class SahteHarita:
            def coz(self, etiket):
                if etiket == "UDOKAN":
                    return {"masraf_merkezi_kodu": "UDOKAN",
                            "masraf_merkezi_adi": "Udokan (GMK)",
                            "sirket": "RHI", "aktif": True}
                return None

        paylasim = [{"masraf_merkezi": "UDOKAN", "pay": 1, "bolen": 2, "oran": 0.5},
                    {"masraf_merkezi": "RHI", "pay": 1, "bolen": 2, "oran": 0.5}]
        t = mahsuplasma_uret([sonuc(gider(tutar=200.0, ek={"paylasim": paylasim}))],
                             SahteHarita())
        merkezler = {m.masraf_merkezi for m in t.satirlar}
        self.assertEqual(merkezler, {"UDOKAN", "GPP"})
        self.assertAlmostEqual(sum(m.tutar for m in t.satirlar), 200.0, places=2)

    def test_bozuk_paylasim_yok_sayilir(self):
        """Oranlarin toplami sifirsa satir bolunmez, tek merkeze yazilir."""
        paylasim = [{"masraf_merkezi": "RHI", "pay": 0, "bolen": 3, "oran": 0.0}]
        t = mahsuplasma_uret([sonuc(gider(tutar=100.0, ek={"paylasim": paylasim}))])
        self.assertEqual(len(t.satirlar), 1)
        self.assertEqual(t.satirlar[0].masraf_merkezi, "GPP")


class RaporlamaTest(unittest.TestCase):
    def test_haritada_olmayan_merkez_isaretlenir(self):
        t = mahsuplasma_uret([sonuc(gider(), merkez="Renservis - Lytkarino",
                                    haritada=False)])
        self.assertFalse(t.satirlar[0].haritada_var)
        self.assertTrue(t.satirlar[0].kontrol_gerek)

    def test_kisi_sayisi_benzersiz(self):
        """Ayni kisi hem otel hem bilet satirinda varsa bir kez sayilmali."""
        s = [sonuc(gider(satir_no=1, gider_tipi="Bilet"), sicil="100001"),
             sonuc(gider(satir_no=2, gider_tipi="Otel", tarih=date(2026, 7, 16)),
                   sicil="100001")]
        t = mahsuplasma_uret(s)
        ozet = t.merkez_ozeti()
        self.assertEqual(len(ozet), 1)
        self.assertEqual(ozet[0]["kisi_sayisi"], 1)
        self.assertEqual(ozet[0]["satir_sayisi"], 2)

    def test_merkez_ozeti_yuzdeleri_yuze_tamamlaniyor(self):
        s = [sonuc(gider(satir_no=1, tutar=300.0), merkez="GPP"),
             sonuc(gider(satir_no=2, tutar=100.0, kisi="MEHMET KAYA"), merkez="AGPP")]
        ozet = mahsuplasma_uret(s).merkez_ozeti()
        self.assertAlmostEqual(sum(k["pay_yuzde"] for k in ozet), 100.0, places=1)

    def test_gider_donemi_araligi(self):
        s = [sonuc(gider(satir_no=1, tarih=date(2026, 5, 3))),
             sonuc(gider(satir_no=2, tarih=date(2026, 7, 28)))]
        t = mahsuplasma_uret(s)
        self.assertEqual(t.satirlar[0].gider_donemi, "05.2026 - 07.2026")

    def test_tek_ay_araliksiz_gosterilir(self):
        t = mahsuplasma_uret([sonuc(gider(tarih=date(2026, 7, 3)))])
        self.assertEqual(t.satirlar[0].gider_donemi, "07.2026")

    def test_durum_sayaclari(self):
        s = [sonuc(gider(satir_no=1), durum=DURUM_OTOMATIK),
             sonuc(gider(satir_no=2, kisi="B KISI"), durum=DURUM_INCELE),
             sonuc(gider(satir_no=3, kisi="C KISI"), durum=DURUM_ESLESMEDI)]
        m = mahsuplasma_uret(s).satirlar[0]
        self.assertEqual((m.otomatik, m.incele, m.eslesmedi), (1, 1, 1))

    def test_bos_girdi_bos_tablo(self):
        t = mahsuplasma_uret([])
        self.assertEqual(t.satirlar, [])
        self.assertEqual(t.kontrol, [])
        self.assertTrue(t.kapali_mi)
        self.assertEqual(t.toplamlar(), {})


class GercekMesajTest(unittest.TestCase):
    """Gercek Temmuz 2026 Outlook mesaji uzerinde uctan uca dogrulama."""

    ANA = Path("ornek_veri/personel/2025_2026_giris_cikis.xlsx")
    YARDIMCI = Path("ornek_veri/personel/1C_Personnel_List_31082026.xlsx")
    MESAJ = Path("ornek_veri/posta/ornek_mail.msg")

    @classmethod
    def setUpClass(cls):
        if not (cls.ANA.is_file() and cls.MESAJ.is_file()):
            raise unittest.SkipTest("ornek veri eksik")
        from masraf.boru import Boru, CalismaAyarlari

        ayarlar = CalismaAyarlari(
            personel_yolu=cls.ANA,
            yardimci_personel_yolu=cls.YARDIMCI if cls.YARDIMCI.is_file() else None,
            veri_dizini="veri", cikti_dizini="cikti",
            defterleri_besle=False, ogrenmeyi_kaydet=False,
        )
        boru = Boru(ayarlar)
        cls.sonuclar = boru.isle([cls.MESAJ])
        cls.tablo = boru.mahsuplasma(cls.sonuclar)

    def test_her_fatura_kapaniyor(self):
        for k in self.tablo.kontrol:
            with self.subTest(fatura=k.kaynak):
                self.assertTrue(
                    k.kapali_mi,
                    f"{k.kaynak}: okunan {k.gelen} != yinelenen {k.yinelenen_tutar} "
                    f"+ dagitilan {k.dagitilan} + dagitilamayan {k.dagitilamayan}",
                )

    def test_seyahat_dosyalari_cift_sayilmiyor(self):
        """Iki dosya ayni 134 islemi tasiyor; net tutar bir dosyaninki olmali."""
        toplam = self.tablo.toplamlar()["USD"]
        self.assertGreater(toplam["gelen"], 100_000)
        self.assertLess(toplam["net"], 60_000,
                        "Seyahat dosyalari cift sayilmis olabilir")
        self.assertGreater(self.tablo.yinelenen_sayisi, 120)

    def test_isaret_celiskisi_bulundu(self):
        """14.07.2026 tarihli 16 USD'lik kalem iki dosyada zit isaretli."""
        self.assertEqual(len(self.tablo.isaret_celiskileri), 1)
        celiski = self.tablo.isaret_celiskileri[0]
        self.assertEqual(celiski.belge_tarihi, date(2026, 7, 14))
        self.assertEqual(sorted(celiski.tutarlar), [-16.0, 16.0])

    def test_kisi_kutukleri_dagilima_girmiyor(self):
        """Katilimci ve saglik listeleri fatura degildir."""
        self.assertEqual(self.tablo.kutuk_satir_sayisi, 100)

    def test_paylasimli_satir_iki_sirkete_bolunuyor(self):
        paylasimlilar = [m for m in self.tablo.satirlar if m.pay_notu]
        self.assertEqual(len(paylasimlilar), 2)
        self.assertAlmostEqual(sum(m.tutar for m in paylasimlilar), 201.42, places=1)

    def test_dagitim_orani_makul(self):
        toplam = self.tablo.toplamlar()["USD"]
        self.assertGreater(toplam["oran"], 90.0)

    def test_excel_mahsuplasma_sayfalari_yaziliyor(self):
        import tempfile

        from masraf.cikti import excel_yaz

        with tempfile.TemporaryDirectory() as gecici:
            yol = Path(gecici) / "test.xlsx"
            excel_yaz(list(self.sonuclar), str(yol), {}, self.tablo)
            self.assertTrue(yol.is_file())
            try:
                import openpyxl
            except ImportError:
                return
            calisma = openpyxl.load_workbook(yol)
            self.assertEqual(calisma.sheetnames[:2], ["Mahsuplasma", "Kontrol"])
            self.assertEqual(calisma["Mahsuplasma"].max_row,
                             len(self.tablo.satirlar) + 2)  # baslik + toplam

    def test_mahsuplasma_csv_yaziliyor(self):
        import csv as _csv
        import tempfile

        from masraf.cikti import mahsuplasma_csv_yaz

        with tempfile.TemporaryDirectory() as gecici:
            yol = Path(gecici) / "mahsup.csv"
            mahsuplasma_csv_yaz(self.tablo, str(yol))
            with open(yol, encoding="utf-8-sig", newline="") as f:
                satirlar = list(_csv.DictReader(f, delimiter=";"))
            self.assertEqual(len(satirlar), len(self.tablo.satirlar))
            self.assertIn("Masraf Merkezi Kodu", satirlar[0])


if __name__ == "__main__":
    unittest.main(verbosity=2)


class ParolaKorumaliTest(unittest.TestCase):
    """Parola korumali dosya anlasilir bir mesajla reddedilmeli.

    Ham istisna metni ('ImportError: msoffcrypto ...') kullaniciya hicbir sey
    anlatmaz. Paketten sifre cozme kutuphanesi cikarildigi icin bu durumun
    net bir mesaj uretmesi gerekir.
    """

    def test_ole_kabindaki_xlsx_korumali_sayilir(self):
        import tempfile

        from masraf.boru import _parola_korumali_mi

        with tempfile.TemporaryDirectory() as gecici:
            # Modern .xlsx bir ZIP'tir; sifrelenince Excel onu OLE kabina koyar.
            sahte = Path(gecici) / "sifreli.xlsx"
            sahte.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 512)
            self.assertTrue(_parola_korumali_mi(sahte))

    def test_normal_dosyalar_korumali_sayilmaz(self):
        """Yanlis pozitif olmamali: saglam dosya reddedilirse is durur."""
        from masraf.boru import _parola_korumali_mi

        for yol in (Path("ornek_veri/personel/2025_2026_giris_cikis.xlsx"),
                    Path("ornek_veri/posta/ornek_mail.msg")):
            if yol.is_file():
                with self.subTest(dosya=yol.name):
                    self.assertFalse(_parola_korumali_mi(yol))

    def test_olmayan_ve_ilgisiz_dosya_false_doner(self):
        import tempfile

        from masraf.boru import _parola_korumali_mi

        self.assertFalse(_parola_korumali_mi(Path("olmayan_dosya.xlsx")))
        with tempfile.TemporaryDirectory() as gecici:
            metin = Path(gecici) / "not.txt"
            metin.write_text("merhaba", encoding="utf-8")
            self.assertFalse(_parola_korumali_mi(metin))
